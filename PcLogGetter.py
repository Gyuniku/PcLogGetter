from datetime import datetime
import dpath
from pytz import timezone
from pymongo import MongoClient
import csv
import openpyxl
import os

GENERAL_COL_LIST = [['ComputerName', 'コンピュータ名'],
                    ['Status.TimeStamp', 'データ取得日時']]

LOGS_COL_LIST = [['Status.Defender.RealTimeProtectionEnabled', 'リアルタイム保護が有効か'],
                 ['Status.Defender.AntivirusEnabled', 'アンチウイルス機能が有効か'],
                 ['Status.Defender.AntispywareEnabled', 'アンチスパイウェア機能が有効か'],
                 ['Status.Defender.AntivirusSignatureVersion', 'ウイルス定義ファイルのバージョン'],
                 ['Status.Defender.AntispywareSignatureVersion', 'スパイウェア定義ファイルのバージョン'],
                 ['Status.Defender.QuickScanEndTime', '最終スキャン完了日時']]

WIN_UPDATE_COL = 'Status.Update'
WIN_UPDATE_COL_AMOUNT = 3
WIN_UPDATE_COL_LIST = [['HotFixID', 'Windows Update HotFixID（最新3件）'],
                       ['InstalledOn', 'インストール日時']]

SOFTWARE_COL = 'Status.Software'
SOFTWARE_COL_LIST = [['DisplayName', 'ソフトウェア名'],
                     ['DisplayVersion', 'バージョン'],
                     ['Publisher', '発行者']]

DATETIME_FORMAT = '%Y/%m/%d %H:%M:%S'
TIMEZONE = 'Asia/Tokyo'

DB_URL = 'mongodb://~~~~~~~~~/'
DB_NAME = 'db_name'
DB_COL_WIN = 'col_name'

XLSX_FILE_NAME = 'result.xlsx'
XLSX_DEFALUT_SHEET = 'Sheet'
# 自動調節するセルの幅に足す定数
XLSX_CELL_WIDTH_CONSTANT = 7
# セルの幅を設定する係数
XLSX_CELL_WIDTH_COEFFICIENT = 1.25


# Windowsのログ用のコレクションに接続するオブジェクトを生成して返します。
def generate_col_win():
    client = MongoClient(DB_URL)
    db = client[DB_NAME]
    return db[DB_COL_WIN]


# DBから取得した辞書型のドキュメントから項目名を使用して必要なデータを抜き出して文字列で返す関数
def fetch_value_from_dic(dic_data, key):
    data = dpath.get(dic_data, key, '.')
    # 日付型の場合はJMTに変換する
    if isinstance(data, datetime):
        data = timezone(TIMEZONE).fromutc(data).strftime(DATETIME_FORMAT)

    return data


# Windows Updateのログデータを最新に並び替え、設定した件数を抜き出して返します。
def extract_win_update(dic_data):
    dic_data.sort(key=lambda x: (x['InstalledOn'] is not None, x['InstalledOn']))
    dic_data.reverse()
    latest = []
    for i in range(0, WIN_UPDATE_COL_AMOUNT):
        latest.append(dic_data[i])
    return latest


# コンピュータ名から、最新のログデータを取得して返します。
def load_latest_log(computer_name):
    return list(col.find({'ComputerName': computer_name}).sort('Status.TimeStamp', -1).limit(1))[0]


# ログデータから、ログ用CSVの行データを返します。
def generate_log_csv_column(log):
    # CSVの1行に書き込むデータ
    csv_column = list()

    # 必要なデータをCSVに追加
    for column in GENERAL_COL_LIST:
        csv_column.append(str(fetch_value_from_dic(log, column[0])))
    for column in LOGS_COL_LIST:
        csv_column.append(str(fetch_value_from_dic(log, column[0])))

    # Windows UpdateのログをCSVに追加
    for win_update in extract_win_update(dpath.get(log, WIN_UPDATE_COL, '.')):
        for column in WIN_UPDATE_COL_LIST:
            csv_column.append(str(fetch_value_from_dic(win_update, column[0])))

    return csv_column


# ログデータから、ソフトウェア一覧CSV用のデータ群を返します。
def generate_softwares_csv_column(log):
    # CSVのデータ
    data = list()

    # ヘッダを追加
    header = list()
    for label in GENERAL_COL_LIST:
        header.append(label[1])
    data.append(header)
    header = list()
    for value in GENERAL_COL_LIST:
        header.append(fetch_value_from_dic(log, value[0]))
    data.append(header)

    # 見やすくするために改行を入れる
    data.append('')

    # ラベルを追加
    labels = list()
    for label in SOFTWARE_COL_LIST:
        labels.append(label[1])
    data.append(labels)

    # ソフトウェア一覧を抜き出す
    for software in dpath.get(log, SOFTWARE_COL, '.'):
        # 行データ
        column = list()

        for value in SOFTWARE_COL_LIST:
            column.append(dpath.get(software, value[0], '.'))

        data.append(column)

    return data


# ログ用CSVのヘッダ項目を返します。
def generate_log_csv_header():
    header = list()

    for label in GENERAL_COL_LIST:
        header.append(label[1])
    for label in LOGS_COL_LIST:
        header.append(label[1])
    for i in range(0, WIN_UPDATE_COL_AMOUNT):
        # Windows Updateのログの数だけ項目名を追加
        header.append(WIN_UPDATE_COL_LIST[0][1])
        header.append(WIN_UPDATE_COL_LIST[1][1])

    return header


# データとファイル名から、CSVを出力します。
def export_to_csv(data, file_name):
    with open(file_name, 'w', encoding='utf-8_sig') as file:
        writer = csv.writer(file, lineterminator='\n')
        writer.writerows(data)


# カレントディレクトリ内のCSVファイル一覧を昇順で返します。
def load_csv_list():
    csv_list = list()

    for file in os.listdir():
        if (os.path.splitext(file)[1]) == '.csv':
            csv_list.append(file)

    csv_list.sort()
    return csv_list


# 新規Excelファイルをカレントディレクトリに作成します。
def create_new_xlsx():
    wb = openpyxl.Workbook()
    wb.save(XLSX_FILE_NAME)


# 指定されたワークシートの幅を自動調節します。
def adjust_cell_width(ws):
    for col in ws.columns:
        # 列で最も長い文字列
        max_length = 0
        # 列名を取得
        column = col[0].column
        for cell in col:
            try:
                # 取得したセルの文字列の長さをチェック
                if len(str(cell.value)) > max_length:
                    # 最も長い場合、最大値を書き換える
                    max_length = len(cell.value)
            except:
                # 空白等のセルは無視
                pass
        # 設定するセルの幅
        adjusted_width = (max_length + XLSX_CELL_WIDTH_CONSTANT) * XLSX_CELL_WIDTH_COEFFICIENT
        ws.column_dimensions[column].width = adjusted_width


# 指定されたCSVファイルを開き、指定されたxlsxファイルの新規シートに追加します。
def csv_to_new_sheet(csv_file_name, xlsx_file_name):
    f = open(csv_file_name, encoding="utf-8_sig")
    reader = csv.reader(f, delimiter=',')

    # xlsxを開いてCSVファイル名と同じシートを生成
    wb = openpyxl.load_workbook(xlsx_file_name)
    ws = wb.create_sheet(title=csv_file_name.split('.')[0])

    # シートに書き込む
    for row in reader:
        ws.append(row)

    f.close()

    # セルの幅を調節
    adjust_cell_width(ws)

    wb.save(xlsx_file_name)


# 指定されたXLSXファイルを開き、指定されたシートを削除します。
def destroy_xlsx_sheet(xlsx_file_name, sheet_name):
    # xlsxを開いてシートを削除
    wb = openpyxl.load_workbook(xlsx_file_name)
    ws = wb[sheet_name]
    wb.remove(ws)
    wb.save(xlsx_file_name)


# DB準備
col = generate_col_win()

# 登録されているコンピュータのリストを取得
computer_list = col.distinct('ComputerName')

# 昇順にする
computer_list.sort()

# ログ用CSVのデータ
csv_columns = list()

# ヘッダを追加
csv_columns.append(generate_log_csv_header())

# 取得したコンピュータごとに処理
for computer in computer_list:
    log = load_latest_log(computer)
    csv_columns.append(generate_log_csv_column(log))
    export_to_csv(generate_softwares_csv_column(log), '(2)ソフトウェア一覧_' + str(computer) + '.csv')

# CSVに出力
export_to_csv(csv_columns, '(1)各PCの情報.csv')

# 新規XLSXファイルを作成
create_new_xlsx()

# 出力したCSV一覧を取得し、XLSXにまとめる
csv_list = load_csv_list()
for csv_file in csv_list:
    csv_to_new_sheet(csv_file, XLSX_FILE_NAME)

# 初期のシートを削除
destroy_xlsx_sheet(XLSX_FILE_NAME, XLSX_DEFALUT_SHEET)

# 使用したCSVファイルを削除
for csv_file in csv_list:
    os.remove(csv_file)
